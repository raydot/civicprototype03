"""
Script to create an Excel template for bulk category creation
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

def create_template():
    """Create Excel template for category bulk upload"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"
    
    # Define headers
    headers = [
        "Name",
        "Description", 
        "Type",
        "Political Spectrum",
        "Policy Areas",
        "Keywords"
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = {
        'A': 30,  # Name
        'B': 50,  # Description
        'C': 15,  # Type
        'D': 20,  # Political Spectrum
        'E': 40,  # Policy Areas
        'F': 60   # Keywords
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add instructions in row 2
    instruction_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    instruction_font = Font(italic=True, size=10, color="856404")
    
    instructions = [
        "Required. Category name",
        "Required. Brief description of the category",
        "Required. Either 'issue' or 'policy'",
        "Required. One of: progressive, leans_left, bipartisan, leans_right, conservative",
        "Required. Comma-separated (e.g., healthcare, economy)",
        "Required. Comma-separated, minimum 3 (e.g., healthcare, insurance, medicare)"
    ]
    
    for col_num, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = instruction
        cell.fill = instruction_fill
        cell.font = instruction_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 60
    
    # Add example rows (rows 3-5)
    examples = [
        {
            "name": "Healthcare Access & Affordability",
            "description": "Universal healthcare, insurance coverage, prescription drug costs, Medicare/Medicaid expansion",
            "type": "issue",
            "spectrum": "progressive",
            "policy_areas": "healthcare, economy, social services",
            "keywords": "healthcare, insurance, medicare, medicaid, prescription drugs, affordable care act, universal healthcare"
        },
        {
            "name": "Tax Policy & Reform",
            "description": "Income tax rates, corporate taxes, wealth taxes, tax credits and deductions",
            "type": "policy",
            "spectrum": "bipartisan",
            "policy_areas": "economy, fiscal policy",
            "keywords": "taxes, income tax, corporate tax, tax reform, tax credits, deductions, IRS"
        },
        {
            "name": "Second Amendment Rights",
            "description": "Gun ownership rights, concealed carry, constitutional interpretation",
            "type": "issue",
            "spectrum": "conservative",
            "policy_areas": "civil rights, public safety",
            "keywords": "second amendment, gun rights, firearms, concealed carry, constitutional rights"
        }
    ]
    
    example_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    example_font = Font(size=10, color="666666", italic=True)
    
    for row_num, example in enumerate(examples, 3):
        cells_data = [
            example["name"],
            example["description"],
            example["type"],
            example["spectrum"],
            example["policy_areas"],
            example["keywords"]
        ]
        
        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = example_fill
            cell.font = example_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
        
        ws.row_dimensions[row_num].height = 50
    
    # Add empty rows for data entry (rows 6-25)
    for row_num in range(6, 26):
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 40
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80
    
    instruction_text = [
        ("BULK CATEGORY UPLOAD TEMPLATE", Font(bold=True, size=16, color="667EEA")),
        ("", None),
        ("HOW TO USE THIS TEMPLATE:", Font(bold=True, size=12)),
        ("", None),
        ("1. Fill in the 'Categories' sheet with your category data", Font(size=11)),
        ("   - Rows 3-5 contain examples (you can delete these)", Font(size=11, italic=True)),
        ("   - Start entering your categories in row 6", Font(size=11, italic=True)),
        ("", None),
        ("2. Required Fields:", Font(bold=True, size=11)),
        ("   • Name: Category name (e.g., 'Healthcare Access')", Font(size=11)),
        ("   • Description: Brief description of what this category covers", Font(size=11)),
        ("   • Type: Must be either 'issue' or 'policy'", Font(size=11)),
        ("   • Political Spectrum: Must be one of:", Font(size=11)),
        ("     - progressive", Font(size=11, italic=True)),
        ("     - leans_left", Font(size=11, italic=True)),
        ("     - bipartisan", Font(size=11, italic=True)),
        ("     - leans_right", Font(size=11, italic=True)),
        ("     - conservative", Font(size=11, italic=True)),
        ("   • Policy Areas: Comma-separated list (e.g., 'healthcare, economy')", Font(size=11)),
        ("   • Keywords: Comma-separated list, minimum 3 required", Font(size=11)),
        ("", None),
        ("3. Save the file and upload it through the Category Admin interface", Font(size=11)),
        ("", None),
        ("4. The system will validate your data and show a preview before creating categories", Font(size=11)),
        ("", None),
        ("IMPORTANT NOTES:", Font(bold=True, size=12, color="DC3545")),
        ("", None),
        ("• This template is for CREATING NEW categories only", Font(size=11)),
        ("• It will not update existing categories", Font(size=11)),
        ("• All fields are required", Font(size=11)),
        ("• Keywords must be comma-separated (e.g., 'keyword1, keyword2, keyword3')", Font(size=11)),
        ("• Minimum 3 keywords per category", Font(size=11)),
        ("• Type must be exactly 'issue' or 'policy' (lowercase)", Font(size=11)),
        ("• Political Spectrum values are case-sensitive (use lowercase with underscores)", Font(size=11)),
        ("", None),
        ("TIPS:", Font(bold=True, size=11)),
        ("", None),
        ("• Use the examples in rows 3-5 as a guide", Font(size=11)),
        ("• Be specific with keywords - they're used to match candidate statements", Font(size=11)),
        ("• Policy areas help organize categories thematically", Font(size=11)),
        ("• Description should clearly explain what the category covers", Font(size=11)),
    ]
    
    for row_num, (text, font) in enumerate(instruction_text, 1):
        cell = ws_instructions.cell(row=row_num, column=1)
        cell.value = text
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_instructions.row_dimensions[row_num].height = 20 if text else 10
    
    # Save template
    output_path = Path(__file__).parent.parent / "app" / "static" / "templates" / "category_upload_template.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    print(f"✅ Template created successfully: {output_path}")
    return output_path

if __name__ == "__main__":
    create_template()
